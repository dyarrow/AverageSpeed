from PyQt5.QtWidgets import QWidget,QMessageBox

import os
import sys
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QMessageBox,QSizePolicy
from PyQt5.QtCore import *
from PyQt5.QtGui import * 
from PyQt5.QtCore import QThread, pyqtSignal, QMutex, QMutexLocker
from datetime import datetime
import time
from PyQt5.QtCore import Qt, QSortFilterProxyModel
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl
from PyQt5.QtWidgets import QFileDialog,QLineEdit, QPushButton, QApplication,QVBoxLayout, QDialog,QAction,QLabel,QTableWidgetItem
from openpyxl import Workbook

import openpyxl
from tablemodel import CustomTableModel,CustomProxyModel,ValidationProxyModel,FilterHeaderView,ColumnFilterDialog
from queue import Queue

import json
import sys

from build_config import resourcesPath
from build_config import applicationPath
from version import BUILD_VERSION, BUILD_DATE

from baseline_measurement import SectionData,ProcessSectionData

from averagespeed_config import AverageSpeedConfig
from link_validation import AverageSpeedLinkValidationManualComparison,AverageSpeedLinkValidationSaveKML
from validation_wizard import ValidationWizard



class ExcelSectionData:
    ID="ID"
    ROADID="RoadID"
    MEASUREMENT_NUMBER="MeasurementNumber"
    SECTION_NUMBER="SectionNumber"
    GPS_FILENAME="GPSFileName"
    TIME_STARTED="TimeStarted"
    TIME_ENDED="TimeEnded"

class ProgressDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter, option, index):
        progress = index.data(QtCore.Qt.UserRole+1000)
        opt = QtWidgets.QStyleOptionProgressBar()
        opt.rect = option.rect
        opt.minimum = 0
        opt.maximum = 100
        opt.progress = progress
        opt.text = "{}%".format(progress)
        opt.textVisible = True
        QtWidgets.QApplication.style().drawControl(QtWidgets.QStyle.CE_ProgressBar, opt, painter)



class SettingsDialog(QtWidgets.QDialog):
    def __init__(self, config, config_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setMinimumWidth(350)
        self.config = config
        self.config_path = config_path

        layout = QtWidgets.QVBoxLayout()

        # GPS Settings
        gps_group = QtWidgets.QGroupBox("GPS")
        gps_form = QtWidgets.QFormLayout()

        self.spin_min_sats = QtWidgets.QSpinBox()
        self.spin_min_sats.setRange(0, 20)
        self.spin_min_sats.setValue(int(float(config["AverageSpeed"]["min_sats"])))
        self.spin_min_sats.setToolTip("Minimum number of GPS satellites required for a point to be included")
        gps_form.addRow("Min Satellites:", self.spin_min_sats)

        self.spin_time_offset = QtWidgets.QDoubleSpinBox()
        self.spin_time_offset.setRange(-24, 24)
        self.spin_time_offset.setDecimals(1)
        self.spin_time_offset.setSingleStep(0.5)
        self.spin_time_offset.setValue(float(config["AverageSpeed"]["time_offset"]))
        self.spin_time_offset.setToolTip("Time offset in hours to apply to GPS timestamps")
        gps_form.addRow("Time Offset (hours):", self.spin_time_offset)

        self.spin_leap_seconds = QtWidgets.QSpinBox()
        self.spin_leap_seconds.setRange(0, 60)
        self.spin_leap_seconds.setValue(int(float(config["AverageSpeed"]["leap_seconds"])))
        self.spin_leap_seconds.setToolTip("GPS leap seconds to subtract from epoch time")
        gps_form.addRow("Leap Seconds:", self.spin_leap_seconds)
        gps_group.setLayout(gps_form)
        layout.addWidget(gps_group)

        # Validation Threshold Settings
        val_group = QtWidgets.QGroupBox("Validation Thresholds")
        val_form = QtWidgets.QFormLayout()

        self.chk_pct_only = QtWidgets.QCheckBox("Percentage Only")
        self.chk_pct_only.setChecked(config["AverageSpeed"].get("pct_only", "false").lower() == "true")
        self.chk_pct_only.setToolTip("Use percentage thresholds for all speeds regardless of speed breakpoint")
        self.chk_pct_only.stateChanged.connect(self._onPctOnlyChanged)
        val_form.addRow("", self.chk_pct_only)

        self.spin_speed_breakpoint = QtWidgets.QDoubleSpinBox()
        self.spin_speed_breakpoint.setRange(0, 200)
        self.spin_speed_breakpoint.setSuffix(" mph")
        self.spin_speed_breakpoint.setValue(float(config["AverageSpeed"].get("speed_breakpoint", 62)))
        self.spin_speed_breakpoint.setToolTip("Speed above which percentage tolerance is used instead of mph")
        val_form.addRow("Speed Breakpoint:", self.spin_speed_breakpoint)

        low_layout = QtWidgets.QHBoxLayout()
        self.spin_threshold_low_pos = QtWidgets.QDoubleSpinBox()
        self.spin_threshold_low_pos.setRange(0, 20); self.spin_threshold_low_pos.setSingleStep(0.5)
        self.spin_threshold_low_pos.setValue(float(config["AverageSpeed"].get("threshold_low_pos", 3)))
        self.spin_threshold_low_pos.setToolTip("Max mph above OBO speed (positive tolerance)")
        self.spin_threshold_low_neg = QtWidgets.QDoubleSpinBox()
        self.spin_threshold_low_neg.setRange(0, 20); self.spin_threshold_low_neg.setSingleStep(0.5)
        self.spin_threshold_low_neg.setValue(float(config["AverageSpeed"].get("threshold_low_neg", 3)))
        self.spin_threshold_low_neg.setToolTip("Max mph below OBO speed (negative tolerance)")
        low_layout.addWidget(QtWidgets.QLabel("+")); low_layout.addWidget(self.spin_threshold_low_pos)
        low_layout.addWidget(QtWidgets.QLabel("−")); low_layout.addWidget(self.spin_threshold_low_neg)
        val_form.addRow("Low Speed Tolerance (mph):", low_layout)

        high_layout = QtWidgets.QHBoxLayout()
        self.spin_threshold_high_pos = QtWidgets.QDoubleSpinBox()
        self.spin_threshold_high_pos.setRange(0, 20); self.spin_threshold_high_pos.setSingleStep(0.5)
        self.spin_threshold_high_pos.setValue(float(config["AverageSpeed"].get("threshold_high_pos", 3)))
        self.spin_threshold_high_pos.setToolTip("Max % above OBO speed (positive tolerance)")
        self.spin_threshold_high_neg = QtWidgets.QDoubleSpinBox()
        self.spin_threshold_high_neg.setRange(0, 20); self.spin_threshold_high_neg.setSingleStep(0.5)
        self.spin_threshold_high_neg.setValue(float(config["AverageSpeed"].get("threshold_high_neg", 3)))
        self.spin_threshold_high_neg.setToolTip("Max % below OBO speed (negative tolerance)")
        high_layout.addWidget(QtWidgets.QLabel("+")); high_layout.addWidget(self.spin_threshold_high_pos)
        high_layout.addWidget(QtWidgets.QLabel("−")); high_layout.addWidget(self.spin_threshold_high_neg)
        val_form.addRow("High Speed Tolerance (%):", high_layout)

        val_group.setLayout(val_form)
        layout.addWidget(val_group)

        # Apply initial enabled state
        self._onPctOnlyChanged()
        layout.addSpacing(8)

        # Buttons
        btn_box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel
        )
        btn_box.accepted.connect(self.save)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.setLayout(layout)

    def _onPctOnlyChanged(self):
        pct_only = self.chk_pct_only.isChecked()
        self.spin_speed_breakpoint.setEnabled(not pct_only)
        self.spin_threshold_low_pos.setEnabled(not pct_only)
        self.spin_threshold_low_neg.setEnabled(not pct_only)

    def save(self):
        self.config["AverageSpeed"]["min_sats"]           = str(self.spin_min_sats.value())
        self.config["AverageSpeed"]["time_offset"]         = str(self.spin_time_offset.value())
        self.config["AverageSpeed"]["leap_seconds"]        = str(self.spin_leap_seconds.value())
        self.config["AverageSpeed"]["pct_only"]            = str(self.chk_pct_only.isChecked()).lower()
        self.config["AverageSpeed"]["speed_breakpoint"]    = str(self.spin_speed_breakpoint.value())
        self.config["AverageSpeed"]["threshold_low_pos"]   = str(self.spin_threshold_low_pos.value())
        self.config["AverageSpeed"]["threshold_low_neg"]   = str(self.spin_threshold_low_neg.value())
        self.config["AverageSpeed"]["threshold_high_pos"]  = str(self.spin_threshold_high_pos.value())
        self.config["AverageSpeed"]["threshold_high_neg"]  = str(self.spin_threshold_high_neg.value())
        try:
            with open(self.config_path, "w") as f:
                json.dump(self.config, f, indent=4)
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Save Failed",
                f"Could not save settings:\n{str(e)}")


#Plugin UI Class
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setWindowTitle("Neology Average Speed Enforcement")
        self.resize(1060, 655)

        try:
            with open(f"{resourcesPath}/neology_average_speed.json") as c:
                self.commissioningConfig = json.load(c)
        except Exception as e:
            print(str(e))
            exit()

        self._build_ui()

        # Restore validation enabled state from config
        val_enabled = self.commissioningConfig.get("AverageSpeed", {}).get("validation_enabled", "false").lower() == "true"
        self.chk_validation_enabled.setChecked(val_enabled)

        # ── Proxy models for Baseline Measurement tab ──────────────────────
        self.tbl_view_section_data_proxy = CustomProxyModel()
        self.tbl_view_section_data_proxy_header = self.tbl_view_section_data.horizontalHeader()
        self.tbl_view_gps_data_proxy = CustomProxyModel()
        self.tbl_view_gps_data_proxy_header = self.tbl_view_gps_data.horizontalHeader()

        # ── State ──────────────────────────────────────────────────────────
        self.section_data = []
        self.progress_bar_value = 0
        self.no_threads = 30
        self.vbox_processing_threads = {}
        self.vbox_processing_objects = {}
        self.thread_progress = {}
        self.vbox = {}
        self.mutex = QMutex()
        self.data = []
        self._last_vrm_groups = []

        self.pbAverageSpeedValidation = None  # replaced by QProgressDialog at runtime
        self.btn_save_kml.setEnabled(False)
        self.btn_export_validation_data.setEnabled(False)
        self.btn_export_vbox_cut.setEnabled(False)

        self.validation_headers = ['Passage','Pri Entry Time','Sec Entry Time','Entry Time Diff','Pri Exit Time','Sec Exit Time','Exit Time Diff','VRM','From','To','Vbox Average Spd','Pri OBO Spd','Vbox/Pri Speed % Diff',"Vbox / Pri Speed MPH Diff","Sec OBO Speed","% Pri/Sec Spd Diff","FromVboxCutTime","ToVboxCutTime",'GPS Points','# Errors (low satellite)']
        self.gps_data_headers = ['Sat #','Epoch','Time','Lat','Long','Speed','Northing','Easting','Altitude']

    def _build_ui(self):
        """Build the entire UI in code — no .ui file required."""

        # ── Stylesheet ─────────────────────────────────────────────────────
        self.setStyleSheet("""
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 9pt;
            }
            QGroupBox {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                margin-top: 8px;
                padding: 4px 6px 4px 6px;
                background: transparent;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 6px;
                color: #555;
                font-size: 8pt;
            }
            QPushButton {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 2px 8px;
                background-color: #ffffff;
                color: #333;
                min-height: 20px;
                font-size: 8.5pt;
            }
            QPushButton:hover {
                background-color: #f0f7ff;
                border-color: #aac4e0;
            }
            QPushButton:pressed {
                background-color: #dceeff;
            }
            QPushButton:disabled {
                color: #bbb;
                border-color: #e5e5e5;
                background-color: #fafafa;
            }
            QProgressBar {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                text-align: center;
                background-color: #f5f5f5;
                height: 16px;
                font-size: 8pt;
                color: #333;
            }
            QProgressBar::chunk {
                background-color: #0078d4;
                border-radius: 3px;
            }
            QTableView {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                gridline-color: #f0f0f0;
                selection-background-color: #cce4f7;
                selection-color: #333;
                alternate-background-color: #fafcff;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                border: none;
                border-right: 1px solid #e0e0e0;
                border-bottom: 1px solid #e0e0e0;
                padding: 4px 8px;
                font-weight: 600;
                color: #444;
            }
            QTabWidget::pane {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                top: -1px;
            }
            QTabBar::tab {
                background: #f5f5f5;
                border: 1px solid #e0e0e0;
                border-bottom: none;
                padding: 5px 16px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                color: #666;
                margin-right: 2px;
                min-width: 140px;
            }
            QTabBar::tab:selected {
                background: #fff;
                color: #0078d4;
                font-weight: 600;
                border-bottom: 2px solid #0078d4;
            }
            QLineEdit {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 3px 6px;
                background: #fff;
                selection-background-color: #cce4f7;
            }
            QLineEdit:focus {
                border-color: #0078d4;
            }
            QDoubleSpinBox, QSpinBox {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 2px 4px;
                background: #fff;
            }
            QDoubleSpinBox:focus, QSpinBox:focus {
                border-color: #0078d4;
            }
            QComboBox {
                border: 1px solid #d0d0d0;
                border-radius: 4px;
                padding: 3px 6px;
                background: #fff;
            }
            QCheckBox {
                spacing: 6px;
                color: #333;
            }
            QCheckBox::indicator {
                width: 14px;
                height: 14px;
                border: 1px solid #d0d0d0;
                border-radius: 3px;
                background: #fff;
            }
            QCheckBox::indicator:checked {
                background-color: #0078d4;
                border-color: #0078d4;
            }
            QFrame[frameShape="5"] {
                color: #d0d0d0;
                max-width: 1px;
                margin: 2px 6px;
            }
        """)

        # ── Central widget & top-level layout ──────────────────────────────
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root_layout = QtWidgets.QVBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # ── Tab widget ─────────────────────────────────────────────────────
        self.tabWidget = QtWidgets.QTabWidget()
        root_layout.addWidget(self.tabWidget)

        # ══ Tab 0: Baseline Measurement (hidden) ═══════════════════════════
        tab0 = QtWidgets.QWidget()
        tab0_layout = QtWidgets.QVBoxLayout(tab0)
        tab0_layout.setContentsMargins(6, 6, 6, 6)
        tab0_layout.setSpacing(4)

        # Button row
        btn_row0 = QtWidgets.QHBoxLayout()
        btn_row0.setSpacing(4)
        self.btn_import_section_data = QtWidgets.QPushButton("Import Section Data")
        self.btn_import_section_data.clicked.connect(self.import_section_data)
        self.btn_process_vbox_data = QtWidgets.QPushButton("Process Vbox Data")
        self.btn_process_vbox_data.clicked.connect(self.import_vbox_data)
        self.btn_export_multiple = QtWidgets.QPushButton("Export Section VBO Files")
        self.btn_export_multiple.clicked.connect(self.export_multiple)
        btn_row0.addWidget(self.btn_import_section_data)
        btn_row0.addWidget(self.btn_process_vbox_data)
        btn_row0.addWidget(self.btn_export_multiple)
        btn_row0.addStretch()
        tab0_layout.addLayout(btn_row0)

        # Section data table
        self.tbl_view_section_data = QtWidgets.QTableView()
        self.tbl_view_section_data.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_view_section_data.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl_view_section_data.horizontalHeader().setStretchLastSection(True)
        self.tbl_view_section_data.verticalHeader().setVisible(False)
        tab0_layout.addWidget(self.tbl_view_section_data)

        # Combo + progress row
        combo_row = QtWidgets.QHBoxLayout()
        self.combo_section_selection = QtWidgets.QComboBox()
        self.combo_section_selection.activated.connect(self.combo_section_selection_changed)
        self.pb_progress = QtWidgets.QProgressBar()
        self.pb_progress.setValue(0)
        combo_row.addWidget(self.combo_section_selection)
        combo_row.addWidget(self.pb_progress)
        tab0_layout.addLayout(combo_row)

        # GPS data table
        self.tbl_view_gps_data = QtWidgets.QTableView()
        self.tbl_view_gps_data.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_view_gps_data.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl_view_gps_data.horizontalHeader().setStretchLastSection(True)
        self.tbl_view_gps_data.verticalHeader().setVisible(False)
        tab0_layout.addWidget(self.tbl_view_gps_data)

        self.tabWidget.addTab(tab0, "Baseline Measurement")
        self.tabWidget.setTabVisible(0, False)

        # ══ Tab 1: Baseline Validation ══════════════════════════════════════
        tab1 = QtWidgets.QWidget()
        tab1_layout = QtWidgets.QVBoxLayout(tab1)
        tab1_layout.setContentsMargins(6, 6, 6, 4)
        tab1_layout.setSpacing(4)

        # ── Toolbar ────────────────────────────────────────────────────────
        _tb_ss = (
            "QToolButton{"
            "background:#ffffff;color:#333;border:1px solid #d0d0d0;"
            "border-radius:4px;padding:3px 10px;font-size:8.5pt;max-height:24px;}"
            "QToolButton:hover{background:#f0f7ff;border-color:#aac4e0;}"
            "QToolButton:pressed{background:#dceeff;}"
            "QToolButton:disabled{color:#bbb;border-color:#e5e5e5;background:#fafafa;}"
            "QToolButton::menu-indicator{width:0;}"
        )

        def _make_menu_btn(label):
            btn = QtWidgets.QToolButton()
            btn.setText(label)
            btn.setPopupMode(QtWidgets.QToolButton.InstantPopup)
            btn.setToolButtonStyle(Qt.ToolButtonTextOnly)
            btn.setStyleSheet(_tb_ss)
            return btn

        toolbar_layout = QtWidgets.QHBoxLayout()
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        toolbar_layout.setSpacing(4)

        # Compare menu
        compare_btn = _make_menu_btn("⚡ Compare ▾")
        compare_menu = QtWidgets.QMenu(compare_btn)
        self.btn_file_check = compare_menu.addAction("Quick Compare")
        self.btn_file_check.triggered.connect(self.btn_file_checkPressed)
        self.btn_wizard = compare_menu.addAction("Wizard")
        self.btn_wizard.triggered.connect(self.btn_wizardPressed)
        compare_btn.setMenu(compare_menu)
        toolbar_layout.addWidget(compare_btn)

        # Export menu
        export_btn = _make_menu_btn("↓ Export ▾")
        export_menu = QtWidgets.QMenu(export_btn)
        self.btn_save_kml = export_menu.addAction("Export KML")
        self.btn_save_kml.triggered.connect(self.btn_save_kmlPressed)
        self.btn_export_validation_data = export_menu.addAction("Export CSV / XLSX")
        self.btn_export_validation_data.triggered.connect(self.export_validation_data)
        self.btn_export_vbox_cut = export_menu.addAction("Export Vbox Cut")
        self.btn_export_vbox_cut.triggered.connect(self.export_vbox_cut_data)
        export_btn.setMenu(export_menu)
        toolbar_layout.addWidget(export_btn)

        # Separator
        sep = QtWidgets.QFrame()
        sep.setFrameShape(QtWidgets.QFrame.VLine)
        sep.setFixedWidth(1)
        toolbar_layout.addWidget(sep)

        # Enable Validation checkbox
        self.chk_validation_enabled = QtWidgets.QCheckBox("Enable Validation")
        self.chk_validation_enabled.stateChanged.connect(self.onValidationEnabledChanged)
        toolbar_layout.addWidget(self.chk_validation_enabled)

        toolbar_layout.addStretch()

        # Settings button
        self.btn_settings = QtWidgets.QPushButton("⚙ Settings")
        self.btn_settings.clicked.connect(self.btnSettingsPressed)
        toolbar_layout.addWidget(self.btn_settings)

        # Help menu
        help_btn = _make_menu_btn("? Help ▾")
        help_menu = QtWidgets.QMenu(help_btn)
        act_open_help = help_menu.addAction("Open Help")
        act_open_help.triggered.connect(self._open_help)
        help_menu.addSeparator()
        act_about = help_menu.addAction("About")
        act_about.triggered.connect(self._show_about)
        help_btn.setMenu(help_menu)
        toolbar_layout.addWidget(help_btn)

        # Hidden import config button (kept for backward compat)
        self.btn_import_config_file = QtWidgets.QPushButton("Import Config")
        self.btn_import_config_file.clicked.connect(self.btnImportConfigFilePressed)
        self.btn_import_config_file.setVisible(False)
        toolbar_layout.addWidget(self.btn_import_config_file)

        tab1_layout.addLayout(toolbar_layout)

        # Validation table
        self.tableValidation = QtWidgets.QTableView()
        self.tableValidation.setFont(QFont("Segoe UI", 8))
        self.tableValidation.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tableValidation.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.tableValidation.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableValidation.setSortingEnabled(True)
        self.tableValidation.horizontalHeader().setMinimumSectionSize(80)
        self.tableValidation.horizontalHeader().setDefaultSectionSize(80)
        self.tableValidation.horizontalHeader().setStretchLastSection(True)
        self.tableValidation.verticalHeader().setVisible(False)
        self.tableValidation.verticalHeader().setMinimumSectionSize(20)
        self.tableValidation.verticalHeader().setDefaultSectionSize(20)
        tab1_layout.addWidget(self.tableValidation)

        self.tabWidget.addTab(tab1, "Baseline Validation")
        self.tabWidget.setCurrentIndex(1)

        # ── FilterHeaderView installed once ────────────────────────────────
        self._filter_header = FilterHeaderView(Qt.Horizontal, self.tableValidation)
        self.tableValidation.setHorizontalHeader(self._filter_header)
        self._filter_header.filter_clicked.connect(self._show_column_filter)
        self._filter_header.setSectionsClickable(True)

        # ── Status bar ─────────────────────────────────────────────────────
        self._status_bar = QtWidgets.QStatusBar()
        self._status_bar.setSizeGripEnabled(False)
        self._status_bar.setStyleSheet(
            "QStatusBar { background: #f5f5f5; border-top: 1px solid #e0e0e0; "
            "font-size: 8pt; color: #555; padding: 0 8px; max-height: 22px; }"
            "QStatusBar::item { border: none; }"
        )
        self.setStatusBar(self._status_bar)
        self._sb_passages = QtWidgets.QLabel("No comparison run")
        self._sb_passed   = QtWidgets.QLabel()
        self._sb_failed   = QtWidgets.QLabel()
        self._sb_time     = QtWidgets.QLabel()
        self._sb_obo      = QtWidgets.QLabel()
        for lbl in [self._sb_passages, self._sb_passed, self._sb_failed, self._sb_time, self._sb_obo]:
            lbl.setStyleSheet("padding: 0 6px;")
        self._status_bar.addWidget(self._sb_passages)
        self._status_bar.addWidget(self._make_sb_sep())
        self._status_bar.addWidget(self._sb_passed)
        self._status_bar.addWidget(self._make_sb_sep())
        self._status_bar.addWidget(self._sb_failed)
        self._status_bar.addPermanentWidget(self._sb_obo)
        self._status_bar.addPermanentWidget(self._make_sb_sep())
        self._status_bar.addPermanentWidget(self._sb_time)

#region Baseline Measurement 
    def import_section_data(self):
        section_data_filename, check = QFileDialog.getOpenFileName(None,"Select Section Data Spreadsheet File:","","Excel File (*.xlsx);;All Files (*.*)",)
        if not check: return


        try:
            self.section_data=[]

            workbook=openpyxl.load_workbook(section_data_filename)
            sheet = workbook['BaselineVboxTimings']

            list_values = list(sheet.values)
            headers=list(list_values[0])

            #Dynamically get excel column numbers (in case someone changes the format of excel file)
            self.id_col=headers.index(ExcelSectionData.ID)
            self.road_id_col=headers.index(ExcelSectionData.ROADID)
            self.measurement_number_col=headers.index(ExcelSectionData.MEASUREMENT_NUMBER)
            self.section_number_col=headers.index(ExcelSectionData.SECTION_NUMBER)
            self.gps_filename_col=headers.index(ExcelSectionData.GPS_FILENAME)
            self.start_time_col=headers.index(ExcelSectionData.TIME_STARTED)
            self.end_time_col=headers.index(ExcelSectionData.TIME_ENDED)

            self.section_data=[list(elem) for elem in list_values[1:]]

            for row in self.section_data:
                if row is not None:
                    row[self.start_time_col]=row[self.start_time_col].strftime("%d/%m/%Y %H:%M:%S.%f").strip()[:-3]
                    row[self.end_time_col]=row[self.end_time_col].strftime("%d/%m/%Y %H:%M:%S.%f").strip()[:-3]
                    
            self.tbl_view_section_data_model = CustomTableModel(self.section_data,headers)      
            
            self.tbl_view_section_data_proxy.setFilterKeyColumn(-1)
            self.tbl_view_section_data_proxy.setSourceModel(self.tbl_view_section_data_model)
            self.tbl_view_section_data.setModel(self.tbl_view_section_data_proxy)
            self.tbl_view_section_data.setSortingEnabled(True)
            
            self.tbl_view_section_data_proxy_header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
            self.tbl_view_section_data_proxy_header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        except Exception as e:
            print(e)
            self.showErrorMessagebox("Section Import Error","Failed to import section data, please check format of excel spreadsheet")

    def import_vbox_data(self):

        self.section_processing_thread=ProcessSectionData(self.section_data,self.id_col,self.gps_filename_col, self.start_time_col, self.end_time_col)
        self.section_processing_thread.section_data_processing_finished.connect(self.on_section_data_processing_finished, QtCore.Qt.QueuedConnection)
        self.section_processing_thread.section_processing_error_occurred.connect(self.on_section_processing_error_occurred, QtCore.Qt.QueuedConnection)
        self.section_processing_thread.update_section_processing_progress.connect(self.update_section_processing_progress, QtCore.Qt.QueuedConnection)
        self.section_processing_thread.start()

        #self.vbox.import_vbox_file("C:\\Users\\dyarrow\\OneDrive - Neology\\Desktop\\WSDOT\\Baseline Measurement\\Trailer speed\\Trailer speed\\SB I-5 all passes.vbo", start_time, end_time)
        #self.vbox.export_to_csv("baseline_tool.csv")

    def update_section_processing_progress(self, progress_dict):
        if "progress" in progress_dict:
            self.pb_progress.setValue(progress_dict["progress"])
            self.pb_progress.setFormat(progress_dict['message'])
            self.pb_progress.setAlignment(QtCore.Qt.AlignCenter)
    
    def combo_section_selection_changed(self):
        self.update_gps_table(self.combo_section_selection.currentText())

    def on_section_data_processing_finished(self, section_data_objects):
        self.section_processing_thread.quit()
        self.section_processing_thread.wait()
        self.update_section_processing_progress({"progress":100,"message":"Finished Processing Section Data"})

        for row in self.section_data:
            #section_data_objects[row[self.id_col]].display_points()
            self.combo_section_selection.addItem(str(row[self.id_col]))

        self.section_data_objects=section_data_objects
        self.update_gps_table()

    def on_section_processing_error_occurred(self, error_message):
        QMessageBox.warning(self, 'Vbox Processing Error', f'Vbox Processing failed with error: {error_message}')

        self.refresh_progress_output()
        
        self.section_processing_thread.quit()
        self.section_processing_thread.wait()

    def update_gps_table(self,section_number=1):
        gps_data_list=self.section_data_objects[int(section_number)].get_gps_point_list()
        
        self.tbl_view_gps_data_model = CustomTableModel(gps_data_list,self.gps_data_headers)      
        
        self.tbl_view_gps_data_proxy.setFilterKeyColumn(-1)
        self.tbl_view_gps_data_proxy.setSourceModel(self.tbl_view_gps_data_model)
        self.tbl_view_gps_data.setModel(self.tbl_view_gps_data_proxy)
        self.tbl_view_gps_data.setSortingEnabled(True)
        
        self.tbl_view_gps_data_proxy_header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.tbl_view_gps_data_proxy_header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)

    def export_multiple(self):
        saveFolder = QtWidgets.QFileDialog.getExistingDirectory(self, 'Select Folder')
        if not saveFolder: return

        if self.section_data_objects != None:
            for section in self.section_data_objects:
                self.section_data_objects[section].export_filtered_vbo(f"{saveFolder}/{self.section_data_objects[section].uid}.vbo")
            self._offer_open(saveFolder)

#endregion

#region Link Validation
    def _row_passes(self, row):
        """Return True if a result row passes the current threshold settings."""
        try:
            cfg      = self.commissioningConfig["AverageSpeed"]
            bp       = float(cfg.get("speed_breakpoint",  62))
            lp       = float(cfg.get("threshold_low_pos",  3))
            ln       = float(cfg.get("threshold_low_neg",  3))
            hp       = float(cfg.get("threshold_high_pos", 3))
            hn       = float(cfg.get("threshold_high_neg", 3))
            pct_only = cfg.get("pct_only", "false").lower() == "true"
            pri_speed = float(row[11])
            pct_diff  = float(row[12])
            mph_diff  = float(row[13])
            if not self.chk_validation_enabled.isChecked():
                return True
            if pct_only:
                return -hn <= pct_diff <= hp
            if pri_speed <= bp:
                return -ln <= mph_diff <= lp
            return -hn <= pct_diff <= hp
        except (ValueError, TypeError, IndexError):
            return True

    def _show_friendly_error(self, raw_error):
        """Show a user-friendly error dialog, translating common exceptions."""
        raw = str(raw_error)

        if "'Matched'" in raw or "Matched" in raw and "KeyError" in raw:
            title = "OBO File Error — Missing Sheet"
            msg   = ("The OBO file does not contain a sheet named <b>Matched</b>.<br><br>"
                     "Please check the OBO export settings and ensure the correct file has been selected.")

        elif "is not in list" in raw or "ValueError" in raw and any(
                col in raw for col in ["Entry Primary Time", "Exit Primary Time",
                                       "Primary Speed", "Secondary Speed",
                                       "LP Number", "LP Hash", "Primary Camera ID"]):
            # Extract the column name if possible
            import re
            match = re.search(r"'([^']+)' is not in list", raw)
            col = match.group(1) if match else "a required column"
            title = "OBO File Error — Missing Column"
            msg   = (f"The OBO file is missing the column: <b>{col}</b><br><br>"
                     "Required columns are:<br>"
                     "<tt>LP Number, LP Hash, Entry Primary Time, Entry Secondary Time,<br>"
                     "Exit Primary Time, Exit Secondary Time, Primary Speed,<br>"
                     "Secondary Speed, Primary Camera ID</tt><br><br>"
                     "Please check the OBO export settings.")

        elif "No passages found" in raw or "no passages" in raw.lower():
            title = "No Passages Found"
            msg   = ("No passages were found in the OBO data matching the entered VRM or hash.<br><br>"
                     "Check that:<br>"
                     "• The VRM or hash is entered correctly<br>"
                     "• The OBO export covers the correct time window<br>"
                     "• The correct OBO file has been selected")

        elif "GPS" in raw or "vbo" in raw.lower() or "csv" in raw.lower():
            title = "GPS File Error"
            msg   = (f"There was a problem reading the GPS file:<br><br>"
                     f"<tt>{raw}</tt><br><br>"
                     "Ensure the file is a valid VBox .vbo or OxTS .csv file.")

        else:
            title = "Comparison Failed"
            msg   = (f"The comparison failed with the following error:<br><br>"
                     f"<tt>{raw}</tt>")

        self.showRichErrorMessagebox(title, msg)

    def _make_sb_sep(self):
        sep = QtWidgets.QFrame()
        sep.setFrameShape(QtWidgets.QFrame.VLine)
        sep.setStyleSheet("color: #d0d0d0; max-width: 1px; margin: 3px 2px;")
        return sep

    def _update_status_bar(self, passages, passed, failed, obo_file=None):
        self._sb_passages.setText(f"Passages: {passages}")
        self._sb_passed.setText(f"✔ Passed: {passed}")
        self._sb_passed.setStyleSheet("padding: 0 6px; color: #107c10; font-weight: 600;")
        self._sb_failed.setText(f"✖ Failed: {failed}")
        self._sb_failed.setStyleSheet(
            "padding: 0 6px; color: #c50f1f; font-weight: 600;"
            if failed > 0 else "padding: 0 6px; color: #107c10; font-weight: 600;"
        )
        import datetime
        self._sb_time.setText(f"Last run: {datetime.datetime.now().strftime('%H:%M:%S')}")
        if obo_file:
            import os
            self._sb_obo.setText(f"OBO: {os.path.basename(obo_file)}")

    def _open_help(self):
        import os, sys
        if getattr(sys, 'frozen', False):
            # Running as built exe — help.html is next to the exe
            base = os.path.dirname(sys.executable)
        else:
            base = applicationPath
        help_path = os.path.join(base, "help.html")
        if not os.path.exists(help_path):
            QtWidgets.QMessageBox.warning(self, "Help Not Found",
                f"Could not find help.html at:\n{help_path}")
            return
        QDesktopServices.openUrl(QUrl.fromLocalFile(help_path))

    def _show_about(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("About Neology Average Speed Enforcement")
        dlg.setFixedWidth(420)
        dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        layout = QtWidgets.QVBoxLayout(dlg)
        layout.setContentsMargins(24, 20, 24, 16)
        layout.setSpacing(10)

        title = QtWidgets.QLabel("Neology Average Speed Enforcement")
        title.setStyleSheet("font-size: 11pt; font-weight: 600;")
        layout.addWidget(title)

        version_lbl = QtWidgets.QLabel(f"Version {BUILD_VERSION}  ·  {BUILD_DATE}")
        version_lbl.setStyleSheet("color: #555; font-size: 8.5pt;")
        layout.addWidget(version_lbl)

        layout.addSpacing(4)

        body_style = "font-size: 8.5pt; color: #333;"
        for text in [
            "Copyright © 2026 Neology Ltd. All rights reserved.",
            "This program uses Qt version 5.15.2 via PyQt5.",
            "Qt is a C++ toolkit for cross-platform application development.",
            "PyQt5 and Qt are licensed under the GNU General Public License (GPL) version 3. "
            "This application is distributed under the terms of the GPL v3. "
            'See <a href="https://gnu.org/licenses/gpl-3.0">gnu.org/licenses/gpl-3.0</a>.',
            'For an overview of Qt licensing, see <a href="https://qt.io/licensing">qt.io/licensing</a>.',
            'Python is used under the Python Software Foundation Licence. '
            'See <a href="https://python.org">python.org</a>.',
        ]:
            lbl = QtWidgets.QLabel(text)
            lbl.setStyleSheet(body_style)
            lbl.setWordWrap(True)
            lbl.setOpenExternalLinks(True)
            layout.addWidget(lbl)

        layout.addSpacing(4)
        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok)
        btn_box.accepted.connect(dlg.accept)
        layout.addWidget(btn_box, alignment=Qt.AlignRight)

        dlg.exec_()

    def btnSettingsPressed(self):
        dlg = SettingsDialog(self.commissioningConfig, f"{resourcesPath}/neology_average_speed.json", self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            try:
                with open(f"{resourcesPath}/neology_average_speed.json") as c:
                    self.commissioningConfig = json.load(c)
                self.recolourValidationTable()
            except Exception as e:
                self.showErrorMessagebox("Settings Error", f"Could not reload settings: {str(e)}")

    def btnImportConfigFilePressed(self):
        xmlFilename, check = QFileDialog.getOpenFileName(None,"Select AverageSpeed EWA Config File:","","XML File (*.xml);;All Files (*.*)",)
        if not check: return
        self.safeZoneConfig=AverageSpeedConfig()
        self.safeZoneConfig.importConfig(xmlFilename)

    def btn_file_checkPressed(self):
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Quick Compare — Vehicle ID")
        dlg.setMinimumWidth(320)
        form = QtWidgets.QFormLayout(dlg)
        form.setContentsMargins(16, 16, 16, 8)
        form.setSpacing(8)
        line_vrm  = QtWidgets.QLineEdit()
        line_vrm.setPlaceholderText("e.g. ANZ6427")
        line_hash = QtWidgets.QLineEdit()
        line_hash.setPlaceholderText("Optional")
        form.addRow("Plate / VRM:", line_vrm)
        form.addRow("Plate Hash:", line_hash)
        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btn_box.accepted.connect(dlg.accept)
        btn_box.rejected.connect(dlg.reject)
        form.addRow(btn_box)

        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return
        vrm        = line_vrm.text().strip()
        plate_hash = line_hash.text().strip()
        if not vrm and not plate_hash:
            self.showErrorMessagebox("Missing Details", "Please enter a Plate or Hash.")
            return

        GPSFilenames, check = QFileDialog.getOpenFileNames(self, "Select GPS File(s)", "", "All Files (*.*);;Vbox Files (*.vbo);;OxTS Files (*.csv)")
        if not check: return
        if any("vbo" in s.lower() for s in GPSFilenames) and any("csv" in s.lower() for s in GPSFilenames):
            self.showErrorMessagebox("Incorrect selection", "Please select only Vbox or only OxTS csv files.")
            return

        OBODataFilenames, check = QFileDialog.getOpenFileNames(self, "Select OBO Data File", "", "OBO Files (*.txt *.xlsx);;Text Files (*.txt);;Excel Files (*.xlsx);;All Files (*.*)")
        if not check: return

        # Detect MDOT format and ask for timezone offset if needed
        obo_offset = self._detect_and_ask_obo_timezone(OBODataFilenames)
        self.commissioningConfig["AverageSpeed"]["obo_time_offset"] = str(obo_offset)

        self._wizard = None
        vrm_groups = [{"plate": vrm, "plate_hash": plate_hash, "gps_files": GPSFilenames}]
        self._startComparison(OBODataFilenames, vrm_groups)

    def btn_wizardPressed(self):
        self._wizard = ValidationWizard(self)
        self._wizard.comparisonRequested.connect(self._onWizardComparisonRequested)
        if self._last_vrm_groups:
            self._wizard.page(self._wizard.VRM_PAGE).set_previous_state(self._last_vrm_groups)
        self._wizard.show()

    def _onWizardComparisonRequested(self):
        # Apply OBO time offset from wizard page to config before comparing
        self.commissioningConfig["AverageSpeed"]["obo_time_offset"] = str(
            self._wizard.get_obo_time_offset()
        )
        self._startComparison(self._wizard.get_obo_files(), self._wizard.get_vrm_groups())

    def _onWizardProgress(self, str_val):
        if self._wizard is None:
            return
        progress = str_val.get("progress", 0)
        message  = str_val.get("message", "")
        page = self._wizard.page(self._wizard.PROGRESS_PAGE)
        page.progress_bar.setValue(progress)
        page.lbl_status.setText(message)

    def _detect_and_ask_obo_timezone(self, filenames):
        """Check if any OBO file is MDOT reduced format. If so ask for timezone offset and return it. Otherwise return 0."""
        try:
            import openpyxl
            from link_validation import MDOTInputData
            for f in filenames:
                if not f.lower().endswith(('.xlsx', '.xls')):
                    continue
                wb = openpyxl.load_workbook(f, read_only=True)
                if 'Matched' not in wb.sheetnames:
                    continue
                rows = list(wb['Matched'].rows)
                if not rows:
                    continue
                headers = [c.value for c in rows[0]]
                if MDOTInputData.detect(headers):
                    return self._ask_obo_timezone_offset()
        except Exception:
            pass
        return 0.0

    def _ask_obo_timezone_offset(self):
        """Show timezone offset dialog and return the chosen offset in hours."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("OBO Timezone Offset")
        dlg.setFixedWidth(340)
        layout = QtWidgets.QVBoxLayout(dlg)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 12)

        lbl = QtWidgets.QLabel(
            "This OBO file uses local timestamps rather than UTC.\n\n"
            "Enter the hours to add to OBO times to convert to UTC:\n"
            "e.g. EDT (UTC\u22124) \u2192 enter +4,  BST (UTC+1) \u2192 enter \u22121"
        )
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        spin = QtWidgets.QDoubleSpinBox()
        spin.setRange(-24, 24)
        spin.setDecimals(1)
        spin.setSingleStep(0.5)
        spin.setValue(float(self.commissioningConfig["AverageSpeed"].get("obo_time_offset", "0")))
        spin.setSuffix(" hours")
        form = QtWidgets.QFormLayout()
        form.addRow("OBO Time Offset:", spin)
        layout.addLayout(form)

        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btn_box.accepted.connect(dlg.accept)
        btn_box.rejected.connect(dlg.reject)
        layout.addWidget(btn_box)

        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            return spin.value()
        return 0.0

    def _startComparison(self, OBODataFilenames, vrm_groups):
        self._last_obo_file = OBODataFilenames[0] if OBODataFilenames else None
        self.btn_save_kml.setEnabled(False)
        self.btn_export_validation_data.setEnabled(False)
        self.btn_export_vbox_cut.setEnabled(False)
        self.btn_file_check.setEnabled(False)
        self.btn_wizard.setEnabled(False)
        self.tableValidation.setModel(None)

        self._pending_groups = list(vrm_groups)
        self._obo_filenames  = OBODataFilenames
        self._group_index    = 0
        self._total_groups   = len(vrm_groups)
        self.linkValidationData = None  # clear previous results
        self._last_progress  = 0       # prevent progress bar going backwards

        # Clear status bar until new results arrive
        self._sb_passages.setText("Running...")
        self._sb_passed.setText("")
        self._sb_failed.setText("")
        self._sb_passed.setStyleSheet("padding: 0 6px;")
        self._sb_failed.setStyleSheet("padding: 0 6px;")

        self._progress_dialog = None
        if not (hasattr(self, '_wizard') and self._wizard is not None):
            self._progress_dialog = QtWidgets.QProgressDialog("Starting comparison...", None, 0, 100, self)
            self._progress_dialog.setWindowTitle("Running Comparison")
            self._progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
            self._progress_dialog.setMinimumWidth(420)
            self._progress_dialog.setCancelButton(None)
            self._progress_dialog.setMinimumDuration(0)
            self._progress_dialog.setValue(0)
            self._progress_dialog.show()

        self._run_next_group()

    def _run_next_group(self):
        if self._group_index >= self._total_groups:
            return

        group      = self._pending_groups[self._group_index]
        plate      = group.get('plate', '')
        plate_hash = group.get('plate_hash', '')
        gps_files  = group.get('gps_files', [])

        self.AverageSpeedValidationManualCheckCompare = AverageSpeedLinkValidationManualComparison(
            gps_files, self._obo_filenames, self.commissioningConfig, plate, plate_hash
        )
        self.AverageSpeedValidationManualCheckCompare.updateAverageSpeedValidationPB.connect(self._on_group_progress)
        self.AverageSpeedValidationManualCheckCompare.updateValidationTable.connect(self.updateValidationTable)
        self.AverageSpeedValidationManualCheckCompare.validationThreadFinished.connect(self._on_group_finished)
        self.AverageSpeedValidationManualCheckCompare.start()

    def _on_group_progress(self, str_val):
        if self._group_index >= self._total_groups:
            return
        msg       = str_val.get('message', '')
        group     = self._pending_groups[self._group_index]
        vrm_label = group.get('plate') or group.get('plate_hash') or f"Group {self._group_index + 1}"

        # Each group owns an equal band of 0–95%
        # Within the band: GPS=25%, OBO=65%, Calculating=90%
        band      = 95.0 / self._total_groups
        band_start = self._group_index * band
        msg_lower = msg.lower()
        if 'gps' in msg_lower:
            within = 0.25
        elif 'ercu' in msg_lower or 'obo' in msg_lower:
            within = 0.65
        elif 'calculat' in msg_lower:
            within = 0.90
        else:
            within = 0.10

        overall = int(band_start + within * band)
        overall = max(overall, self._last_progress)
        self._last_progress = overall

        if self._total_groups > 1:
            label = f"[{self._group_index + 1}/{self._total_groups}] [{vrm_label}] {msg}"
        else:
            label = f"[{vrm_label}] {msg}"

        scaled = {'progress': overall, 'message': label}
        self.updateAverageSpeedValidationPB(scaled)
        if hasattr(self, '_wizard') and self._wizard is not None:
            self._onWizardProgress(scaled)

    def _on_group_finished(self, result):
        if result['Result']:
            self._group_index += 1
            if self._group_index < self._total_groups:
                self._run_next_group()
                return
        self.validationThreadFinished(result)



    def _offer_open(self, filepath):
        reply = QtWidgets.QMessageBox.question(
            self, "Open File",
            f"File saved successfully.\nWould you like to open it?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.Yes
        )
        if reply == QtWidgets.QMessageBox.Yes:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(filepath)
            elif sys.platform == "darwin":
                subprocess.call(["open", filepath])
            else:
                subprocess.call(["xdg-open", filepath])

    def btn_save_kmlPressed(self):
        saveFilename, check = QFileDialog.getSaveFileName(None,"Export Google Earth KML File:","","Google Earth (*.kml);;All Files (*.*)",)
        if not check: return

        self._kml_save_filename = saveFilename
        self.AverageSpeedLinkValidationSaveKML=AverageSpeedLinkValidationSaveKML(saveFilename,self.linkValidationData,self.commissioningConfig)
        self.AverageSpeedLinkValidationSaveKML.updateAverageSpeedValidationPB.connect(self.updateAverageSpeedValidationPB)
        self.AverageSpeedLinkValidationSaveKML.validationThreadFinished.connect(self.kmlThreadFinished)
        self.AverageSpeedLinkValidationSaveKML.start()

    def kmlThreadFinished(self, result):
        if result['Result']:
            self._offer_open(self._kml_save_filename)
        else:
            self.showErrorMessagebox(result['Title'], result['Text'])

    def export_validation_data(self):
        saveFilename, check = QFileDialog.getSaveFileName(None,"Save Comparison Data:","","CSV (*.csv);;All Files (*.*)",)
        if not check: return

        with open(saveFilename,"w") as outfile:
            outfile.write(','.join(map(str, self.validation_headers)) + "\n")

            for row in self.linkValidationData.validationResultData:
                outfile.write(','.join(map(str, row)) + "\n")
            
        wb = Workbook()
        ws = wb.active
        ws.append(self.validation_headers)
        for row in self.linkValidationData.validationResultData:
            ws.append(row)
        wb.save(f"{saveFilename}.xlsx")
        self._offer_open(saveFilename)


    def export_vbox_cut_data(self):
        # Derive a default filename from the VRMs in the results
        try:
            vrms = "_".join(dict.fromkeys(
                str(row[7]) for row in self.linkValidationData.validationResultData
            ))
        except Exception:
            vrms = "vbox"
        saveFilename, check = QFileDialog.getSaveFileName(None, "Save Vbox Cut Data:", f"{vrms}_vbox_cut_data.csv", "CSV Files (*.csv);;All Files (*.*)")
        if not check: return

        with open(saveFilename, "w") as f:
            f.write("PassageID,Sats,Time,Speed,Lat,Long\n")
            for row in self.linkValidationData.vboxCutData:
                f.write(",".join(str(v) for v in row) + "\n")
        self._offer_open(saveFilename)

    def updateValidationTable(self, linkValidationData):
        # Accumulate results across VRM groups — don't replace, append
        if not hasattr(self, 'linkValidationData') or self.linkValidationData is None or self._group_index == 0:
            self.linkValidationData = linkValidationData
        else:
            self.linkValidationData.validationResultData.extend(linkValidationData.validationResultData)
            self.linkValidationData.ercuData.extend(linkValidationData.ercuData)
            self.linkValidationData.gpsData.extend(linkValidationData.gpsData)

        if len(self.linkValidationData.validationResultData):
            self.model = CustomTableModel(self.linkValidationData.validationResultData, self.validation_headers)
            self.proxy_model = ValidationProxyModel()
            self.proxy_model.setSourceModel(self.model)
            self.proxy_model.sort(0, Qt.AscendingOrder)
            self.tableValidation.setModel(self.proxy_model)
            self._filter_header.setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
            self._filter_header.resizeSections(QtWidgets.QHeaderView.ResizeToContents)
            self._filter_header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
            self.recolourValidationTable()

    def setValidationControlsEnabled(self, enabled):
        pass

    def onValidationEnabledChanged(self):
        self.setValidationControlsEnabled(self.chk_validation_enabled.isChecked())
        self.recolourValidationTable()
        # Persist state to config
        self.commissioningConfig.setdefault("AverageSpeed", {})["validation_enabled"] = \
            "true" if self.chk_validation_enabled.isChecked() else "false"
        try:
            import json
            with open(f"{resourcesPath}/neology_average_speed.json", "w") as f:
                json.dump(self.commissioningConfig, f, indent=4)
        except Exception:
            pass

    def recolourValidationTable(self):
        # Vbox/Pri Speed % Diff is col 12, Vbox/Pri Speed MPH Diff is col 13
        # Vbox Average Spd is col 10, Pri OBO Spd is col 11
        if not hasattr(self, 'model') or self.model is None:
            return

        if not self.chk_validation_enabled.isChecked():
            for row in range(self.model.rowCount(None)):
                for col in range(len(self.validation_headers)):
                    self.model.change_color(row, col, None)
            self._sb_passages.setText(f"Passages: {self.model.rowCount(None)}")
            self._sb_passed.setText("")
            self._sb_failed.setText("")
            return

        cfg = self.commissioningConfig["AverageSpeed"]
        low_pos_mph      = float(cfg.get("threshold_low_pos",  3))
        low_neg_mph      = float(cfg.get("threshold_low_neg",  3))
        high_pos_pct     = float(cfg.get("threshold_high_pos", 3))
        high_neg_pct     = float(cfg.get("threshold_high_neg", 3))
        speed_breakpoint = float(cfg.get("speed_breakpoint",  62))
        pct_only         = cfg.get("pct_only", "false").lower() == "true"

        GREEN = QColor(144, 238, 144)
        RED   = QColor(255, 102, 102)

        for row in range(self.model.rowCount(None)):
            try:
                pri_speed  = float(self.model._data[row][11])   # Pri OBO Speed
                mph_diff   = float(self.model._data[row][13])   # Vbox/Pri Speed MPH Diff
                pct_diff   = float(self.model._data[row][12])   # Vbox/Pri Speed % Diff

                if pct_only:
                    passed = (-high_neg_pct <= pct_diff <= high_pos_pct)
                elif pri_speed <= speed_breakpoint:
                    # mph_diff = vbox - obo, so positive means vbox is faster
                    passed = (-low_neg_mph <= mph_diff <= low_pos_mph)
                else:
                    passed = (-high_neg_pct <= pct_diff <= high_pos_pct)

                colour = GREEN if passed else RED
                for col in range(len(self.validation_headers)):
                    self.model.change_color(row, col, colour)
            except (ValueError, TypeError, IndexError):
                pass

        # Refresh status bar to reflect updated pass/fail counts
        total  = self.model.rowCount(None)
        passed = sum(1 for row in self.linkValidationData.validationResultData
                     if self._row_passes(row))
        self._update_status_bar(total, passed, total - passed,
                                getattr(self, '_last_obo_file', None))

    def _show_column_filter_at_pos(self, pos):
        header = self.tableValidation.horizontalHeader()
        col = header.logicalIndexAt(pos)
        if col >= 0:
            self._show_column_filter(col)

    def _show_column_filter(self, col):
        if not hasattr(self, 'model') or self.model is None:
            return
        unique_vals = set()
        for row in self.model._data:
            try:
                unique_vals.add(row[col])
            except IndexError:
                pass
        header_name = self.validation_headers[col] if col < len(self.validation_headers) else str(col)
        dlg = ColumnFilterDialog(col, header_name, self.proxy_model, unique_vals, self)
        dlg.move(QCursor.pos())
        dlg.exec_()
        # Update header text and funnel icons
        active_indices = set()
        for i, name in enumerate(self.validation_headers):
            if self.proxy_model.has_filter(i):
                self.model.setHeaderData(i, Qt.Horizontal, f"{name} ▼")
                active_indices.add(i)
            else:
                self.model.setHeaderData(i, Qt.Horizontal, name)
        self._filter_header.set_filtered_columns(active_indices)

    def updateAverageSpeedValidationPB(self, str_val):
        if hasattr(self, '_progress_dialog') and self._progress_dialog is not None:
            if "progress" in str_val:
                self._progress_dialog.setValue(str_val["progress"])
            self._progress_dialog.setLabelText(str_val["message"])

 
    def validationThreadFinished(self, result):
        if hasattr(self, '_progress_dialog') and self._progress_dialog is not None:
            self._progress_dialog.close()
            self._progress_dialog = None
        self.btn_file_check.setEnabled(True)
        self.btn_wizard.setEnabled(True)
        if result['Result']:
            total = len(self.linkValidationData.validationResultData) if self.linkValidationData else 0

            if total:
                self.btn_save_kml.setEnabled(True)
                self.btn_export_validation_data.setEnabled(True)
                self.btn_export_vbox_cut.setEnabled(True)

                # Build pass/fail counts using current threshold settings and show in wizard
                if hasattr(self, '_wizard') and self._wizard is not None:
                    total  = len(self.linkValidationData.validationResultData)
                    passed = 0
                    failed = 0
                    cfg = self.commissioningConfig["AverageSpeed"]
                    bp  = float(cfg.get("speed_breakpoint",  62))
                    lp  = float(cfg.get("threshold_low_pos",  3))
                    ln  = float(cfg.get("threshold_low_neg",  3))
                    hp  = float(cfg.get("threshold_high_pos", 3))
                    hn  = float(cfg.get("threshold_high_neg", 3))
                    pct_only = self.commissioningConfig["AverageSpeed"].get("pct_only", "false").lower() == "true"
                    val_on   = self.chk_validation_enabled.isChecked()
                    for row in self.linkValidationData.validationResultData:
                        try:
                            pri_speed = float(row[11])
                            mph_diff  = float(row[13])
                            pct_diff  = float(row[12])
                            if not val_on:
                                passed += 1
                            elif pct_only:
                                passed += 1 if (-hn <= pct_diff <= hp) else 0
                            elif pri_speed <= bp:
                                passed += 1 if (-ln <= mph_diff <= lp) else 0
                            else:
                                passed += 1 if (-hn <= pct_diff <= hp) else 0
                        except (ValueError, TypeError, IndexError):
                            passed += 1
                    failed = total - passed
                    self._wizard.show_results(total, passed, failed, self.commissioningConfig,
                                              validation_enabled=self.chk_validation_enabled.isChecked())
                    self._last_vrm_groups = self._wizard.get_vrm_groups()
                    self._wizard.mark_complete()
                    self._wizard.show()

            # Update status bar regardless of passage count
            import datetime as _dt
            if total:
                if self.chk_validation_enabled.isChecked():
                    passed = sum(1 for row in self.linkValidationData.validationResultData
                                 if self._row_passes(row))
                    self._update_status_bar(total, passed, total - passed,
                                            getattr(self, '_last_obo_file', None))
                else:
                    self._sb_passages.setText(f"Passages: {total}")
                    self._sb_passed.setText("")
                    self._sb_failed.setText("")
                    self._sb_time.setText(f"Last run: {_dt.datetime.now().strftime('%H:%M:%S')}")
            else:
                self._sb_passages.setText("Passages: 0")
                self._sb_passed.setText("")
                self._sb_failed.setText("")
                self._sb_time.setText(f"Last run: {_dt.datetime.now().strftime('%H:%M:%S')}")

            if "DisplayMessage" in result:
                self.showInfoMessagebox(result['Title'], result['Text'])

            # Reduced format warning always comes first
            if getattr(self.linkValidationData, 'secondary_data_missing', False):
                QtWidgets.QMessageBox.warning(
                    self, "Reduced OBO Format Detected",
                    "The imported OBO file uses a reduced format that does not include:\n\n"
                    "  • Secondary Entry / Exit Times\n"
                    "  • Entry / Exit Time Differences\n"
                    "  • Secondary Speed\n"
                    "  • Primary / Secondary Speed Difference\n"
                    "  • Camera ID\n\n"
                    "These columns will be blank in the results table."
                )

            # Only warn about no passages if not a reduced format timezone issue
            if total == 0 and not getattr(self.linkValidationData, 'secondary_data_missing', False):
                QtWidgets.QMessageBox.information(
                    self, "No Passages Found",
                    "The comparison completed but no matching passages were found.\n\n"
                    "Check that:\n"
                    "  • The VRM or hash is entered correctly\n"
                    "  • The OBO export covers the correct time window\n"
                    "  • The GPS files cover the same time period as the OBO data"
                )
        else:
            self.btn_save_kml.setEnabled(False)
            self.btn_export_validation_data.setEnabled(False)
            self._show_friendly_error(result.get('Text', ''))


    def showRichErrorMessagebox(self, title, maintext):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle(title)
        msg.setTextFormat(Qt.RichText)
        msg.setText(maintext)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def showInfoMessagebox(self,title, maintext):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(maintext)
        msg.setWindowTitle(title)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
    
    def showErrorMessagebox(self,title, maintext):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText(maintext)
        msg.setWindowTitle(title)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
#endregion


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())